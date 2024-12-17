import json
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.utils.dateparse import parse_datetime

import datetime
import openpyxl
from django.http import HttpResponse
from django.db.models import Count

from .models import Robot

@csrf_exempt
def robot_create(request):
    
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            model = data.get('model')
            version = data.get('version')
            created = data.get('created')

            if not all([model, version, created]):
                return JsonResponse({"error": "All fields (model, version, created) are required."}, status=400)

            if len(model) != 2:
                return JsonResponse({"error": "Model must have exactly 2 characters."}, status=400)
            if len(version) != 2:
                return JsonResponse({"error": "Version must have exactly 2 characters."}, status=400)


            created_date = parse_datetime(created)
            if not created_date:
                return JsonResponse({"error": "Invalid date format."}, status=400)


            robot = Robot.objects.create(
                model=model,
                version=version,
                created=created_date
            )
            return JsonResponse({"message": "Robot created successfully.", "id": robot.id}, status=201)

        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON format."}, status=400)

    return JsonResponse({"error": "Only POST method is allowed."}, status=405)



def generate_production_report(request):

    workbook = openpyxl.Workbook()
    

    end_date = datetime.date.today()
    start_date = end_date - datetime.timedelta(days=7)
    
    data = Robot.objects.filter(
        created__date__range=(start_date, end_date)
    ).values('model', 'version').annotate(
        weekly_count=Count('id')
    ).order_by('model', 'version')
    
    
    models = set(item['model'] for item in data)
    for model in models:

        worksheet = workbook.create_sheet(title=model)
        worksheet.append(["Модель", "Версия", "Количество за неделю"])  
        

        for item in data:
            if item['model'] == model:
                worksheet.append([item['model'], item['version'], item['weekly_count']])
    

    if 'Sheet' in workbook.sheetnames:
        workbook.remove(workbook['Sheet'])
    

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="production_report.xlsx"'
    workbook.save(response)
    return response